from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.database import database
from app.models import StationImportRow, calc_aqi


EXPECTED_COLUMNS = {
    "code",
    "name",
    "ward_id",
    "ward_code",
    "lat",
    "lng",
    "pm25",
    "pm10",
    "traffic_level",
    "construction",
    "factory_near",
    "note",
}


def _normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "co"}


def _normalize_row(raw_row: dict[str, Any]) -> StationImportRow:
    normalized = {key.strip().lower(): _normalize_value(value) for key, value in raw_row.items()}
    payload = {
        "code": normalized.get("code"),
        "name": normalized.get("name"),
        "ward_id": int(normalized["ward_id"]) if normalized.get("ward_id") is not None else None,
        "ward_code": normalized.get("ward_code"),
        "lat": float(normalized["lat"]),
        "lng": float(normalized["lng"]),
        "pm25": float(normalized["pm25"]) if normalized.get("pm25") is not None else None,
        "pm10": float(normalized["pm10"]) if normalized.get("pm10") is not None else None,
        "traffic_level": int(normalized["traffic_level"]) if normalized.get("traffic_level") is not None else 0,
        "construction": _to_bool(normalized.get("construction")),
        "factory_near": float(normalized["factory_near"]) if normalized.get("factory_near") is not None else 0.0,
        "note": normalized.get("note"),
    }
    return StationImportRow(**payload)


async def parse_station_upload(file: UploadFile) -> list[StationImportRow]:
    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File rong")

    rows: list[dict[str, Any]] = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raise HTTPException(status_code=400, detail="File Excel rong")
        headers = [str(item).strip().lower() if item is not None else "" for item in values[0]]
        for row_values in values[1:]:
            if not any(item is not None and str(item).strip() for item in row_values):
                continue
            row = {headers[index]: row_values[index] if index < len(row_values) else None for index in range(len(headers))}
            rows.append(row)
    else:
        raise HTTPException(status_code=400, detail="Chi ho tro file CSV hoac XLSX")

    if not rows:
        raise HTTPException(status_code=400, detail="Khong co dong du lieu hop le")

    missing_required = {"code", "name", "lat", "lng"} - {str(key).strip().lower() for key in rows[0].keys()}
    if missing_required:
        raise HTTPException(
            status_code=400,
            detail=f"Thieu cot bat buoc: {', '.join(sorted(missing_required))}",
        )

    try:
        return [_normalize_row(row) for row in rows]
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Khong doc duoc file import: {exc}") from exc


async def _resolve_ward_id(row: StationImportRow) -> int:
    if row.ward_id is not None:
        ward = await database.fetch_one("SELECT id FROM wards WHERE id = :id", {"id": row.ward_id})
        if ward:
            return int(ward["id"])
    if row.ward_code:
        ward = await database.fetch_one("SELECT id FROM wards WHERE code = :code", {"code": row.ward_code})
        if ward:
            return int(ward["id"])
    raise HTTPException(status_code=400, detail=f"Khong xac dinh duoc phuong cho tram {row.code}")


async def import_station_rows(rows: list[StationImportRow], current_user: dict) -> dict:
    created = 0
    updated = 0
    processed_codes: list[str] = []

    for row in rows:
        ward_id = await _resolve_ward_id(row)
        aqi = calc_aqi(row.pm25, row.pm10)
        existing = await database.fetch_one("SELECT id FROM stations WHERE code = :code", {"code": row.code})
        params = {
            "code": row.code,
            "name": row.name,
            "ward_id": ward_id,
            "lat": row.lat,
            "lng": row.lng,
            "pm25": row.pm25,
            "pm10": row.pm10,
            "aqi": aqi,
            "traffic_level": row.traffic_level,
            "construction": row.construction,
            "factory_near": row.factory_near,
            "note": row.note,
        }
        if existing:
            await database.execute(
                """
                UPDATE stations
                SET name=:name, ward_id=:ward_id, lat=:lat, lng=:lng, pm25=:pm25, pm10=:pm10,
                    aqi=:aqi, traffic_level=:traffic_level, construction=:construction,
                    factory_near=:factory_near, note=:note, timestamp=NOW()
                WHERE code=:code
                """,
                params,
            )
            station_id = int(existing["id"])
            updated += 1
        else:
            inserted = await database.fetch_one(
                """
                INSERT INTO stations (
                    code, name, ward_id, lat, lng, pm25, pm10, aqi,
                    traffic_level, construction, factory_near, note
                )
                VALUES (
                    :code, :name, :ward_id, :lat, :lng, :pm25, :pm10, :aqi,
                    :traffic_level, :construction, :factory_near, :note
                )
                RETURNING id
                """,
                params,
            )
            station_id = int(inserted["id"])
            created += 1

        await database.execute(
            """
            INSERT INTO readings (station_id, pm25, pm10, aqi)
            VALUES (:sid, :pm25, :pm10, :aqi)
            """,
            {"sid": station_id, "pm25": row.pm25, "pm10": row.pm10, "aqi": aqi},
        )
        processed_codes.append(row.code)

    return {
        "created": created,
        "updated": updated,
        "total_rows": len(rows),
        "processed_codes": processed_codes,
        "actor": current_user["username"],
    }
